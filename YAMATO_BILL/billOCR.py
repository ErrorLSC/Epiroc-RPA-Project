#from Faxreader import table_exporter
from pdfspliter import splitPDF
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import openpyxl

def AzureOCR(document_path,endpoint,credential):
    # set `<your-endpoint>` and `<your-key>` variables with the values from the Azure portal
    #endpoint = "https://takanoocr.cognitiveservices.azure.com/"
    credentialkey = AzureKeyCredential(credential)

    document_analysis_client = DocumentAnalysisClient(endpoint, credentialkey)

    #document_path = r"C:\Users\jpeqz\Desktop\Troubles\IT\OrderAcknowledgement.PDF"

    #with io.open(document_path, "rb") as fd:
    #    document = fd.read()

    poller = document_analysis_client.begin_analyze_document("prebuilt-layout", document_path)
    result = poller.result()
    return result

def OCR(pdffile):
    page_list = splitPDF(pdffile)
    result_list = []
    for page in page_list:
        result = AzureOCR(page,endpoint,credential)
        result_list.append(result)
    table_exporter(result_list)
    return

def table_exporter(result_list):
    wb = openpyxl.Workbook()
    i = 1
    for page in result_list:
        if len(page.tables) > 0:

       
            for table_idx, table in enumerate(page.tables):
                wb.create_sheet(title="page"+ str(i)+"table"+str(table_idx+1))
                sheet = wb["page"+ str(i)+"table"+str(table_idx+1)]
                #sheet.cell(row=1,column=1,value="table"+str(table_idx+1))
                for cell in table.cells:
                    sheet.cell(row=cell.row_index+1,column=cell.column_index+1,value=cell.content)
                #table_row_length = table_row_length + table.row_count + 3
        i = i + 1
    emptysheet = wb["Sheet"]
    wb.remove(emptysheet)
    wb.save("Yamatotable.xlsx")

if __name__ == '__main__':
    pdffile = r"C:\Users\jpeqz\OneDrive - Epiroc\SCX\Bills\Yamato\202404\【4月度】御請求書.pdf"
    endpoint = "https://faxformrecoginzer.cognitiveservices.azure.com/"
    credential = "9ab7cca139354de19484d263e5b4e603"
    OCR(pdffile)