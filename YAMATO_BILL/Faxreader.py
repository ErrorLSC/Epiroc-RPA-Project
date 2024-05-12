# -*- coding: utf-8 -*-
"""
Created on Thu Jan 19 10:02:39 2023

@author: JPEQZ
"""

import re
import pandas as pd
import numpy as np
from AzureOCR import AzureOCR 
from pdfspliter import splitPDF
import openpyxl


# 字符串处理函数

def itemnumberreader(content):
    
    content = content.replace(" ","")
    content = content.replace("-","")
    item_number_pattern = re.compile("\d{8,}",re.S)
    item_number_list = re.findall(item_number_pattern,content)
    
    # print("Item numberは以下です:\n",format(item_number_list))
    
    return item_number_list

def address_scanner(content):
    # 市町村まで
    address_pattern = re.compile("((.{2,3}?[都道府県])((?:高市|静岡市|野々市|西村|西八代郡市|芳賀郡市|羽村|福岡市|神戸市|神崎郡市|相模原市|田村|熊本市|浜松市|武蔵村|横浜市|柴田郡村|杵島郡大町|東村|札幌市|新潟市|廿日市|広島市|川崎市|岡山市|大阪市|大町|大村|堺市|四日市|名古屋市中村|名古屋市|吉野郡下市|千葉市|十日町|北村|北九州市|余市郡余市|余市|佐波郡玉村|仙台市|京都市|中新川郡上市|さいたま市|.)?.*?[市区町村])([^1-9一二三四五六七八九十]+))")
    address_list = re.findall(address_pattern,content)
    #print(address_list)
    if len(address_list) > 0:
        for i in range(len(address_list)):
            address_list[i] = address_list[i][0]
            address_list[i] = address_list[i].replace(" ","")
            print("住所を検出しました")
            print("----------------------------------------------------")
            print(address_list)
            print("----------------------------------------------------")
    else: print("住所を検出できませんでした")
    
    return address_list

def customer_number_scanner_byaddress(address_book,address_list):  
    addressbook = pd.read_excel(address_book)
    addressbook['RKADD'].replace(np.nan,"",regex=True,inplace=True)
    if len(address_list) > 0:
        print("顧客番号は以下から選べます")
        for address in address_list: 
            result = addressbook[addressbook['RKADD'].str.contains(address,regex=True)]
            if len(result) > 0:
                print(result)
    else: print("顧客番号は見つかりませんでした。")
    #print(result)        
    
def excel_exporter(result_list):
    wb = openpyxl.Workbook()
    i = 1
    for page in result_list:
        wb.create_sheet(title="page"+ str(i))
        sheet = wb["page"+ str(i)]
        i = i + 1
        # OCR内容储存
        sheet.cell(row=1,column=1,value=page.content)

        # 件号输出
        item_list = itemnumberreader(page.content)
        if len(item_list) > 0:
            sheet.cell(row=3,column=1,value="部品番号検出しました")
            for j in range(len(item_list)):
                sheet.cell(row=4+j,column=1,value=item_list[j])
        else:
            sheet.cell(row=3,column=1,value="部品番号検出できませんでした")

        # OCR表格内容输出
        if len(page.tables) > 0:
            table_row_length = len(item_list) + 3
            # k = 0
            for table_idx, table in enumerate(page.tables):

                sheet.cell(row=table_row_length + 2,column=1,value="table"+str(table_idx+1))
                for cell in table.cells:
                    sheet.cell(row=table_row_length+ 3 +cell.row_index,column=cell.column_index+1,value=cell.content)
                table_row_length = table_row_length + table.row_count + 3
    emptysheet = wb["Sheet"]
    wb.remove(emptysheet)
    wb.save("OCRresult.xlsx")

def table_exporter(result_list):
    wb = openpyxl.Workbook()
    i = 1
    for page in result_list:
        if len(page.tables) > 0:

        # OCR内容储存
        #sheet.cell(row=1,column=1,value=page.content)

        # 件号输出
        # item_list = itemnumberreader(page.content)
        # if len(item_list) > 0:
        #     sheet.cell(row=3,column=1,value="部品番号検出しました")
        #     for j in range(len(item_list)):
        #         sheet.cell(row=4+j,column=1,value=item_list[j])
        # else:
        #     sheet.cell(row=3,column=1,value="部品番号検出できませんでした")

        # OCR表格内容输出
        
            # table_row_length = len(item_list) + 3
            # k = 0
            for table_idx, table in enumerate(page.tables):
                wb.create_sheet(title="page"+ str(i)+"table"+str(table_idx+1))
                sheet = wb["page"+ str(i)+"table"+str(table_idx+1)]
                #sheet.cell(row=1,column=1,value="table"+str(table_idx+1))
                for cell in table.cells:
                    sheet.cell(row=cell.row_index+1,column=cell.column_index+1,value=cell.content)
                #table_row_length = table_row_length + table.row_count + 3
        i = i + 1
    emptysheet = wb["Sheet"]
    wb.remove(emptysheet)
    wb.save("Yamatotable.xlsx")
# def exptopandas(result_list):
#     for page in result_list:
#         if len(page.tables) > 0:
#             ocrlist = []
#             for table_idx, table in enumerate(page.tables):
#                 for i in range(table.column_count):
#                     exec('table{}column{}=[]'.format(table_idx,i))
#                 dynamic_variable = locals()
#                 #dynamic_variable['table{}column{}'.format(table_idx,cell.column_index)] = []
#                 for cell in table.cells:                    
#                     #ocr = "table{}column{}.append('{}')".format(table_idx,table.column_count,cell.content)
#                     dynamic_variable['table{}column{}'.format(table_idx,cell.column_index)].append(cell.content)
#                     ocrlist.append(dynamic_variable['table{}column{}'.format(table_idx,cell.column_index)])
#                     print(ocrlist)
    

if __name__ == '__main__':

    address_book = r"C:\Code\workingdics\OCR\customerlist.xlsx"    
    endpoint = "https://faxformrecoginzer.cognitiveservices.azure.com/"
    credential = "9ab7cca139354de19484d263e5b4e603"
    document = r""  

    page_list = splitPDF(document)
    result_list = []
    for page in page_list:
        result = AzureOCR(page,endpoint,credential)
        result_list.append(result)

    for result in result_list:
        #itemnumberreader(result.content)
        #address_list = address_scanner(result.content)
        #customer_number_scanner_byaddress(address_book,address_list)
        exptopandas(result_list)